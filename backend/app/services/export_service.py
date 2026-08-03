import csv
import io
import os
from datetime import date

from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font
from fpdf import FPDF

from app.models.transaction import Transaction

COLUMNS = ["Date", "Description", "Category", "Type", "Amount", "Notes"]

_FONT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "assets", "fonts")
_FONT_REGULAR = os.path.join(_FONT_DIR, "DejaVuSans.ttf")
_FONT_BOLD = os.path.join(_FONT_DIR, "DejaVuSans-Bold.ttf")


def _make_pdf() -> FPDF:
    """
    Core PDF fonts (Helvetica etc.) only support latin-1, which blows up
    on anything from an em-dash to Indonesian currency formatting edge
    cases to emoji in a transaction description. DejaVu Sans is bundled
    with the app (app/assets/fonts) so this doesn't depend on whatever
    fonts happen to be installed on the host/container.
    """

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_font("DejaVu", "", _FONT_REGULAR)
    pdf.add_font("DejaVu", "B", _FONT_BOLD)
    pdf.add_page()

    return pdf


class ExportService:

    @staticmethod
    def _get_transactions(
        db: Session,
        user_id: int,
        date_from: date | None,
        date_to: date | None
    ) -> list[Transaction]:

        query = db.query(Transaction).filter(Transaction.user_id == user_id)

        if date_from:
            query = query.filter(Transaction.date >= date_from)

        if date_to:
            query = query.filter(Transaction.date <= date_to)

        return query.order_by(Transaction.date).all()

    @staticmethod
    def to_csv(db: Session, user_id: int, date_from=None, date_to=None) -> bytes:

        transactions = ExportService._get_transactions(db, user_id, date_from, date_to)

        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(COLUMNS)

        for t in transactions:
            writer.writerow([
                t.date.isoformat(), t.description, t.category,
                t.type, t.amount, t.notes or ""
            ])

        return buffer.getvalue().encode("utf-8-sig")  # BOM so Excel opens UTF-8 cleanly

    @staticmethod
    def to_excel(db: Session, user_id: int, date_from=None, date_to=None) -> bytes:

        transactions = ExportService._get_transactions(db, user_id, date_from, date_to)

        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"

        ws.append(COLUMNS)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for t in transactions:
            ws.append([
                t.date.isoformat(), t.description, t.category,
                t.type, t.amount, t.notes or ""
            ])

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None) if any(
                cell.value is not None for cell in column_cells
            ) else 10
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 40)

        buffer = io.BytesIO()
        wb.save(buffer)

        return buffer.getvalue()

    @staticmethod
    def to_pdf(db: Session, user_id: int, date_from=None, date_to=None) -> bytes:

        transactions = ExportService._get_transactions(db, user_id, date_from, date_to)

        pdf = _make_pdf()
        pdf.set_font("DejaVu", "B", 14)
        pdf.cell(0, 10, "ExpenseTrackerAI - Transactions Export", ln=True)

        pdf.set_font("DejaVu", "", 9)
        range_label = ""

        if date_from or date_to:
            range_label = f"Period: {date_from or '...'} to {date_to or '...'}"
        pdf.cell(0, 6, range_label, ln=True)
        pdf.ln(2)

        col_widths = [25, 80, 40, 25, 30, 77]

        pdf.set_font("DejaVu", "B", 10)
        for header, width in zip(COLUMNS, col_widths):
            pdf.cell(width, 8, header, border=1)
        pdf.ln()

        pdf.set_font("DejaVu", "", 9)
        total_income = 0.0
        total_expense = 0.0

        for t in transactions:

            row = [
                t.date.isoformat(),
                (t.description or "")[:45],
                t.category,
                t.type,
                f"{t.amount:,.0f}",
                (t.notes or "")[:45],
            ]

            for value, width in zip(row, col_widths):
                pdf.cell(width, 7, str(value), border=1)
            pdf.ln()

            if t.type == "income":
                total_income += t.amount
            else:
                total_expense += t.amount

        pdf.ln(3)
        pdf.set_font("DejaVu", "B", 10)
        pdf.cell(0, 8, f"Total income: {total_income:,.0f}   |   Total expense: {total_expense:,.0f}   |   Balance: {total_income - total_expense:,.0f}", ln=True)

        return bytes(pdf.output())
